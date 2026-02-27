import streamlit as st
import pandas as pd
from io import BytesIO
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image

# ==========================================================
# CONFIGURAÇÃO INICIAL
# ==========================================================
st.set_page_config(page_title="Gestão COC Anestesia", layout="wide")
st.title("Registro de Cirurgias e Dashboard Financeiro")

# ==========================================================
# CONEXÃO GOOGLE SHEETS
# ==========================================================
try:
    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    credentials = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scope
    )

    client = gspread.authorize(credentials)

    # Coloque o link de edição correto da sua planilha
    URL_PLANILHA = "https://docs.google.com/spreadsheets/d/15J8J3dCVdGmn5GY8CWXef-JrpkxHcdX2CErKgWS0q7k/edit"
    planilha = client.open_by_url(URL_PLANILHA)

    aba_cirurgias = planilha.worksheet("CIRURGIAS")
    aba_convenios = planilha.worksheet("Página2")
    aba_cbhpm = planilha.worksheet("Página3")

except gspread.SpreadsheetNotFound:
    st.error("❌ Planilha não encontrada. Verifique o ID e se o service account tem acesso.")
    st.stop()
except Exception as e:
    st.error(f"Erro de conexão com as planilhas: {e}")
    st.stop()

# ==========================================================
# CACHE DE DADOS
# ==========================================================
@st.cache_data(ttl=300)
def carregar_dados():
    df_cirurgias = pd.DataFrame(aba_cirurgias.get_all_records())
    df_convenios = pd.DataFrame(aba_convenios.get_all_records())
    df_cbhpm = pd.DataFrame(aba_cbhpm.get_all_records())
    return df_cirurgias, df_convenios, df_cbhpm

df_cirurgias, df_convenios, df_cbhpm = carregar_dados()

# ==========================================================
# FUNÇÕES AUXILIARES
# ==========================================================
def limpar_moeda(valor):
    if pd.isna(valor) or str(valor).strip() in ['-', '', '0']:
        return 0.0
    valor_str = str(valor).replace('R$', '').replace('.', '').replace(',', '.').strip()
    try:
        return float(valor_str)
    except:
        return 0.0

def formatar_real(valor):
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def converter_para_horas(duracao_str):
    try:
        texto = str(duracao_str).strip()
        if not texto or texto == "nan":
            return None
        partes = texto.split(':')
        if len(partes) != 2:
            return None
        return int(partes[0]) + (int(partes[1]) / 60.0)
    except:
        return None

mapa_cbhpm = df_cbhpm.set_index('Código').to_dict('index') if not df_cbhpm.empty else {}
mapa_convenios = df_convenios.set_index('Convênio').to_dict('index') if not df_convenios.empty else {}

def calcular_faturamento_memoria(row):
    convenio = str(row.get('CONVÊNIO', '')).strip()
    procs_str = str(row.get('PROCEDIMENTO', '')).strip()
    if not convenio or not procs_str or procs_str == 'nan':
        return 0.0

    linha_convenio = mapa_convenios.get(convenio)
    if not linha_convenio:
        return 0.0

    valor_total = 0.0
    lista_procs = procs_str.split('\n')

    for i, proc in enumerate(lista_procs):
        codigo = proc.split(" - ")[0].strip()
        linha_cbhpm = mapa_cbhpm.get(codigo)
        if not linha_cbhpm:
            continue

        porte = str(linha_cbhpm.get('Porte Anest.', '')).strip()
        preco = 0.0
        if porte.isdigit():
            col_an = f"AN{porte}"
            if col_an in linha_convenio:
                preco = limpar_moeda(linha_convenio[col_an])

        if i == 0:
            valor_total += preco
        else:
            valor_total += preco * 0.5
    return valor_total

# ==========================================================
# DASHBOARD
# ==========================================================
st.subheader("🏆 Ranking de Rentabilidade")

if df_cirurgias.empty:
    st.info("Nenhuma cirurgia registrada.")
    st.stop()

df_cirurgias['Valor Virtual'] = df_cirurgias.apply(calcular_faturamento_memoria, axis=1)

if 'DURAÇÃO' not in df_cirurgias.columns:
    st.error("Coluna DURAÇÃO não encontrada.")
    st.stop()

df_cirurgias['Horas'] = df_cirurgias['DURAÇÃO'].apply(converter_para_horas)
df_cirurgias['R$/Hora'] = df_cirurgias.apply(
    lambda row: row['Valor Virtual'] / row['Horas']
    if row['Horas'] is not None and row['Horas'] > 0 else None,
    axis=1
)

faturamento_total = df_cirurgias['Valor Virtual'].sum()
total_cirurgias = len(df_cirurgias)
ticket_medio = faturamento_total / total_cirurgias if total_cirurgias else 0

col1, col2, col3 = st.columns(3)
col1.metric("💰 Faturamento Total", formatar_real(faturamento_total))
col2.metric("🏥 Nº Cirurgias", total_cirurgias)
col3.metric("📊 Ticket Médio", formatar_real(ticket_medio))

df_validos = df_cirurgias.dropna(subset=['R$/Hora']).copy()

if not df_validos.empty:
    st.subheader("📊 Ranking por R$/Hora")
    df_ranking = df_validos.sort_values(by='R$/Hora', ascending=False)
    st.dataframe(df_ranking, use_container_width=True)

    # ==========================================================
    # GERAR EXCEL ESTILIZADO
    # ==========================================================
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_ranking.to_excel(writer, index=False, sheet_name='Ranking')
    buffer.seek(0)

    wb = openpyxl.load_workbook(buffer)
    ws = wb['Ranking']

    # Inserir logo no canto superior esquerdo
    try:
        logo = Image("logo_coc.png")  # imagem na mesma pasta do app
        logo.width = 150
        logo.height = 50
        ws.add_image(logo, 'A1')
    except Exception as e:
        st.warning(f"Não foi possível inserir a imagem do logo: {e}")

    # Aplicar cor condicional na coluna "R$/Hora"
    col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() == 'r$/hora':
            col_idx = idx
            break

    if col_idx:
        cor_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cor_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                valor = float(cell.value)
            except (ValueError, TypeError):
                valor = None
            if valor is not None:
                if valor >= 100:  # limite ajustável
                    cell.fill = cor_verde
                else:
                    cell.fill = cor_vermelho

    # Salvar Excel estilizado
    new_buffer = BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)

    st.download_button(
        "⬇️ Baixar Relatório Excel com Estilo",
        data=new_buffer,
        file_name="Relatorio_COC_Anestesia_Estilizado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ==========================================================
    # ANÁLISE E GRÁFICOS
    # ==========================================================
    st.subheader("🏥 Análise Estratégica por Convênio")
    resumo_convenio = df_validos.groupby('CONVÊNIO').agg({'Valor Virtual':'sum','Horas':'sum'})
    resumo_convenio['R$/Hora'] = resumo_convenio['Valor Virtual'] / resumo_convenio['Horas']
    resumo_convenio['% Faturamento'] = resumo_convenio['Valor Virtual'] / faturamento_total * 100 if faturamento_total>0 else 0
    resumo_convenio = resumo_convenio.sort_values(by='Valor Virtual', ascending=False)

    c1, c2, c3 = st.columns(3)
    c1.metric("🥇 Maior Faturamento", resumo_convenio.index[0], formatar_real(resumo_convenio.iloc[0]['Valor Virtual']))
    c2.metric("⚡ Mais Rentável (R$/Hora)",
              resumo_convenio.sort_values(by='R$/Hora', ascending=False).index[0],
              formatar_real(resumo_convenio.sort_values(by='R$/Hora', ascending=False).iloc[0]['R$/Hora']))
    c3.metric("📉 Menos Rentável (R$/Hora)",
              resumo_convenio.sort_values(by='R$/Hora').index[0],
              formatar_real(resumo_convenio.sort_values(by='R$/Hora').iloc[0]['R$/Hora']))

    st.subheader("📊 Faturamento por Convênio")
    st.bar_chart(resumo_convenio['Valor Virtual'])

    st.subheader("⚡ Rentabilidade por Hora")
    st.bar_chart(resumo_convenio['R$/Hora'])

    resumo_exibicao = resumo_convenio.copy()
    resumo_exibicao['Valor Virtual'] = resumo_exibicao['Valor Virtual'].apply(formatar_real)
    resumo_exibicao['R$/Hora'] = resumo_exibicao['R$/Hora'].apply(formatar_real)
    resumo_exibicao['% Faturamento'] = resumo_exibicao['% Faturamento'].apply(lambda x: f"{x:.1f}%")

    st.subheader("📋 Tabela Estratégica Completa")
    st.dataframe(resumo_exibicao, use_container_width=True)